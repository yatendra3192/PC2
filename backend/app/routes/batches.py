from __future__ import annotations
import json
import logging
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException

from app.db.client import fetch_all, fetch_one, execute_returning, execute
from app.auth.dependencies import get_current_user, require_admin
from app.models.user import TokenPayload
from app.models.batch import BatchResponse
from app.storage.client import save_upload, get_upload_path

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("", response_model=BatchResponse)
async def upload_batch(
    file: UploadFile = File(...),
    client_id: str = Form(...),
    async_process: bool = Form(True),
    user: TokenPayload = Depends(require_admin),
):
    """Upload a file, parse it into product records."""
    # Determine file type
    ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename else "unknown"
    file_type_map = {"pdf": "pdf", "csv": "csv", "xlsx": "xlsx", "xls": "xlsx", "jpg": "image", "png": "image", "jpeg": "image"}
    file_type = file_type_map.get(ext, "csv")

    # Save file
    contents = await file.read()
    storage_path = await save_upload(contents, file.filename or "upload", client_id)

    # Get supplier template for this client (if exists)
    supplier_template = await fetch_one(
        "SELECT id FROM supplier_templates WHERE is_active = true ORDER BY created_at DESC LIMIT 1",
    )

    # Create batch record
    row = await execute_returning(
        """INSERT INTO batches (client_id, supplier_template_id, file_name, file_path, file_type, status, created_by)
           VALUES ($1::uuid, $2::uuid, $3, $4, $5, 'queued', $6::uuid)
           RETURNING *""",
        client_id,
        str(supplier_template["id"]) if supplier_template else None,
        file.filename, storage_path, file_type, user.sub,
    )
    batch_id = str(row["id"])

    # Get pipeline config for client
    client = await fetch_one("SELECT pipeline_config_id FROM clients WHERE id = $1::uuid", client_id)
    pipeline_config_id = str(client["pipeline_config_id"]) if client and client["pipeline_config_id"] else None

    # Parse the file to extract product rows
    records = []
    try:
        file_full_path = get_upload_path(storage_path)
        if file_type == "csv":
            import csv as csv_mod
            with open(file_full_path, "r", encoding="utf-8-sig") as f:
                reader = csv_mod.DictReader(f)
                for i, r in enumerate(reader):
                    r["_row_number"] = i + 1
                    records.append(r)
        elif file_type == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(file_full_path, read_only=True)
            ws = wb.active
            headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for i, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                r = {h: str(v) if v is not None else "" for h, v in zip(headers, row_data)}
                r["_row_number"] = i + 1
                records.append(r)
        logger.info(f"Parsed {len(records)} records from {file_full_path}")
    except Exception as parse_err:
        logger.warning(f"Could not parse file, creating 1 demo product: {parse_err}")

    # If no records parsed (PDF, image, or parse failure), create one demo product
    if not records:
        records = [{"Product Name": "Orbit 24V 6-Zone Smart Irrigation Controller", "Model": "B-0624W", "_row_number": 1}]

    # Create a product for each record
    product_ids = []
    for record in records:
        product_name = record.get("Product Name") or record.get("product_name") or record.get("Name")
        model_number = record.get("Model") or record.get("model_number") or record.get("Model Number")
        sku = record.get("SKU") or record.get("sku")
        brand = record.get("Brand") or record.get("brand")
        supplier = record.get("Supplier") or record.get("supplier_name")

        prod_row = await execute_returning(
            """INSERT INTO products
               (client_id, batch_id, product_name, model_number, sku, brand, supplier_name,
                status, pipeline_config_id)
               VALUES ($1::uuid, $2::uuid, $3, $4, $5, $6, $7, 'draft', $8::uuid)
               RETURNING id""",
            client_id, batch_id, product_name, model_number, sku, brand, supplier,
            pipeline_config_id,
        )
        product_id = str(prod_row["id"])
        product_ids.append(product_id)

        # Store raw field values
        for field_name, value in record.items():
            if field_name.startswith("_"):
                continue
            await execute_returning(
                """INSERT INTO product_raw_values
                   (product_id, supplier_field_name, raw_value, source, source_cell_ref)
                   VALUES ($1::uuid, $2, $3, 'csv', $4)
                   RETURNING id""",
                product_id, field_name, str(value),
                f"row:{record.get('_row_number', 0)}",
            )

    # Update batch with counts
    await execute(
        "UPDATE batches SET item_count = $1, processed_count = 0 WHERE id = $2::uuid",
        len(product_ids), batch_id,
    )

    return _batch_from_row(row)


@router.get("", response_model=list[BatchResponse])
async def list_batches(
    status: str | None = None,
    user: TokenPayload = Depends(get_current_user),
):
    if status:
        rows = await fetch_all(
            "SELECT * FROM batches WHERE status = $1 ORDER BY created_at DESC LIMIT 50", status,
        )
    else:
        rows = await fetch_all("SELECT * FROM batches ORDER BY created_at DESC LIMIT 50")
    return [_batch_from_row(r) for r in rows]


@router.get("/{batch_id}")
async def get_batch(batch_id: str, user: TokenPayload = Depends(get_current_user)):
    """Get batch with all product items and their pipeline status."""
    row = await fetch_one("SELECT * FROM batches WHERE id = $1::uuid", batch_id)
    if not row:
        raise HTTPException(status_code=404, detail="Batch not found")

    products = await fetch_all(
        """SELECT p.id, p.product_name, p.model_number, p.current_stage, p.status,
                  p.completeness_pct, p.overall_confidence,
                  (SELECT COUNT(*) FROM product_iksula_values piv
                   WHERE piv.product_id = p.id AND piv.review_status IN ('needs_review','low_confidence')) as review_count
           FROM products p
           WHERE p.batch_id = $1::uuid
           ORDER BY p.created_at""",
        batch_id,
    )

    batch = _batch_from_row(row)
    return {
        **batch.model_dump(),
        "products": [
            {
                "id": str(p["id"]),
                "product_name": p["product_name"],
                "model_number": p["model_number"],
                "current_stage": p["current_stage"],
                "status": p["status"],
                "completeness_pct": p["completeness_pct"],
                "overall_confidence": p["overall_confidence"],
                "review_count": p["review_count"],
            }
            for p in products
        ],
    }


@router.post("/{batch_id}/process")
async def process_batch(batch_id: str, user: TokenPayload = Depends(require_admin)):
    """Run the full pipeline on all products in a batch."""
    # Ensure all stage processors are registered
    import app.pipeline.stage_1_ingest  # noqa: F401
    import app.pipeline.stage_2_classify  # noqa: F401
    import app.pipeline.stage_3_dedup  # noqa: F401
    import app.pipeline.stage_4_enrich  # noqa: F401
    import app.pipeline.stage_5_validate  # noqa: F401
    import app.pipeline.stage_6_transform  # noqa: F401
    import app.pipeline.stage_7_review  # noqa: F401
    from app.pipeline.orchestrator import run_pipeline

    batch = await fetch_one("SELECT * FROM batches WHERE id = $1::uuid", batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    await execute("UPDATE batches SET status = 'processing' WHERE id = $1::uuid", batch_id)

    products = await fetch_all("SELECT id FROM products WHERE batch_id = $1::uuid", batch_id)

    results = []
    for prod in products:
        product_id = str(prod["id"])
        try:
            result = await run_pipeline(product_id)
            product = await fetch_one("SELECT current_stage, status FROM products WHERE id = $1::uuid", product_id)
            results.append({
                "product_id": product_id,
                "current_stage": product["current_stage"] if product else 1,
                "status": product["status"] if product else "draft",
            })
        except Exception as e:
            logger.error(f"Pipeline failed for product {product_id}: {e}")
            results.append({"product_id": product_id, "status": "failed", "error": str(e)})

    await execute(
        "UPDATE batches SET status = 'complete', processed_count = $1, completed_at = now() WHERE id = $2::uuid",
        len(products), batch_id,
    )

    return {"batch_id": batch_id, "products_processed": len(results), "results": results}


def _batch_from_row(r) -> BatchResponse:
    return BatchResponse(
        id=str(r["id"]),
        client_id=str(r["client_id"]),
        file_name=r["file_name"],
        file_path=r["file_path"],
        file_type=r["file_type"],
        item_count=r["item_count"],
        processed_count=r["processed_count"] or 0,
        status=r["status"],
        error_message=r["error_message"],
        created_by=str(r["created_by"]) if r["created_by"] else None,
        created_at=r["created_at"],
        completed_at=r["completed_at"],
    )
